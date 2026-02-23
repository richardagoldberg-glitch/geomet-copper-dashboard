#!/usr/bin/env python3
"""
Geomet Recycling - Copper Intelligence Dashboard v4.1
Fix Window, Fed Funds, Warehouse Stocks, Price Context, DXY, S/R
"""

import json, os, csv, glob, re, time, hashlib, secrets, calendar
from datetime import datetime, timedelta
from pathlib import Path
from http.server import HTTPServer, SimpleHTTPRequestHandler
from http.cookies import SimpleCookie
import socketserver

# Load .env file into environment before config
_env_path = Path(__file__).parent / ".env"
if _env_path.exists():
    with open(_env_path) as _f:
        for _line in _f:
            _line = _line.strip()
            if _line and not _line.startswith("#") and "=" in _line:
                _k, _, _v = _line.partition("=")
                os.environ.setdefault(_k.strip(), _v.strip())

DATA_DIR = Path(__file__).parent / "data"
POSITION_CSV = DATA_DIR / "geomet_position.csv"
SPREAD_HISTORY = DATA_DIR / "spread_history.json"
STATIC_DIR = Path(__file__).parent / "static"
PORT = 8777

import importlib.util
def load_config():
    cfg_path = Path(__file__).parent / "config.py"
    defaults = {
        "METALS_DEV_API_KEY": "", "LME_MANUAL_USD_MT": 0,
        "FIX_TARGET": 5.90, "GTC_LEVELS": [5.90, 5.95, 6.00, 6.05],
        "TRUCKLOAD_LBS": 42000, "BASELINE_LBS": 200000,
        "ATTENTION_MOVE": 0.10, "BIG_MOVE": 0.20,
        "COMEX_WAREHOUSE_MT": 0, "COMEX_WAREHOUSE_DATE": "",
        "COMEX_WAREHOUSE_TREND": "", "FRED_API_KEY": "",
        "FED_FUNDS_RATE": "4.25-4.50", "FED_FUNDS_MIDPOINT": 4.375,
        "MONTHLY_FLOW": {"Chops": 171800, "BB": 162700, "#2": 106600, "#1": 82700},
    }
    if cfg_path.exists():
        try:
            spec = importlib.util.spec_from_file_location("config", cfg_path)
            mod = importlib.util.module_from_spec(spec)
            spec.loader.exec_module(mod)
            for k, v in defaults.items():
                defaults[k] = getattr(mod, k, v)
        except Exception as e:
            print(f"[WARN] config.py error: {e}")
    return defaults

CFG = load_config()
MT_TO_LB = 2204.62


# ---------------------------------------------------------------------------
# LME CACHE — 30 min cache + market hours only
# ---------------------------------------------------------------------------
_lme_cache = {"price_mt": None, "price_lb": None, "timestamp": 0, "source": "none"}

def fetch_lme_price():
    global _lme_cache
    now = time.time()
    _now = datetime.now()
    _hour = _now.hour; _wd = _now.weekday()
    _lme_open = (
        (_wd == 6 and _hour >= 19) or
        (_wd in (0, 1, 2, 3)) or
        (_wd == 4 and _hour < 13)
    )
    if not _lme_open and _lme_cache["price_lb"]:
        return _lme_cache
    if _lme_cache["price_lb"] and (now - _lme_cache["timestamp"]) < 1800:
        return _lme_cache

    api_key = CFG["METALS_DEV_API_KEY"]
    if api_key:
        try:
            import urllib.request
            url = f"https://api.metals.dev/v1/metal/spot?api_key={api_key}&metal=copper&currency=USD"
            req = urllib.request.Request(url, headers={"User-Agent": "GeometDashboard/1.0"})
            with urllib.request.urlopen(req, timeout=10) as resp:
                data = json.loads(resp.read().decode())
                print(f"[DEBUG] metals.dev response keys: {list(data.keys())}")
                print(f"[DEBUG] metals.dev full response: {json.dumps(data, default=str)[:800]}")
                if "rate" in data:
                    rate = data["rate"]
                    price_mt = rate["price"] if isinstance(rate, dict) else rate
                    price_lb = round(price_mt / MT_TO_LB, 4)
                    # metals.dev /v1/metal/spot returns ~LME 3M (confirmed vs CQG LDKZA)
                    _lme_cache = {"price_mt": round(price_mt, 2), "price_lb": price_lb,
                                  "timestamp": now, "source": "metals.dev"}
                    return _lme_cache
        except Exception as e:
            print(f"[WARN] metals.dev API error: {e}")

    manual = CFG["LME_MANUAL_USD_MT"]
    if manual and manual > 0:
        price_lb = round(manual / MT_TO_LB, 4)
        _lme_cache = {"price_mt": manual, "price_lb": price_lb, "timestamp": now, "source": "manual"}
        return _lme_cache
    return {"price_mt": None, "price_lb": None, "timestamp": now, "source": "none"}


# ---------------------------------------------------------------------------
# REAL-TIME PRICE — investing.com API + yfinance intraday fallback
# ---------------------------------------------------------------------------
_rt_cache = {"price": None, "prev_close": None, "timestamp": 0, "source": None}
_prev_settle_cache = {}  # keyed by ticker: {"price": ..., "timestamp": ...}

def _fetch_prev_settle_yf(ticker="HG=F"):
    """Get previous session's settlement from yfinance daily data (5-min cache per ticker)."""
    global _prev_settle_cache
    now = time.time()
    cached = _prev_settle_cache.get(ticker, {})
    if cached.get("price") and (now - cached.get("timestamp", 0)) < 300:
        return cached["price"]
    try:
        import yfinance as yf
        t = yf.Ticker(ticker)
        hd = t.history(period="5d", interval="1d")
        if not hd.empty and len(hd) >= 2:
            hd = hd.reset_index()
            hd.columns = [c if isinstance(c, str) else c[0] for c in hd.columns]
            for i in range(len(hd)):
                print(f"[DEBUG] yf daily {ticker} bar {i}: {hd.iloc[i]['Date']} close={hd.iloc[i]['Close']:.4f}")
            today_date = datetime.now().date()
            last_date = hd.iloc[-1]["Date"]
            if hasattr(last_date, 'date'):
                last_date = last_date.date()
            elif hasattr(last_date, 'to_pydatetime'):
                last_date = last_date.to_pydatetime().date()
            if last_date >= today_date:
                prev = round(float(hd.iloc[-2]["Close"]), 4)
                print(f"[INFO] Prev settle {ticker} (today in data): ${prev:.4f} from {hd.iloc[-2]['Date']}")
            else:
                prev = round(float(hd.iloc[-1]["Close"]), 4)
                print(f"[INFO] Prev settle {ticker} (no today): ${prev:.4f} from {hd.iloc[-1]['Date']}")
            _prev_settle_cache[ticker] = {"price": prev, "timestamp": now}
            return prev
    except Exception as e:
        print(f"[WARN] Prev settle fetch error ({ticker}): {e}")
        import traceback; traceback.print_exc()
    return None


def _get_active_yf_ticker():
    """Determine which COMEX copper contract to show based on FND proximity.
    Returns (ticker, active_contract, label) where active_contract is 'front' or 'next'.
    Near FND (<=5 trading days), show the next month as liquidity migrates.
    """
    today = datetime.now().date()
    MONTHS = [
        ("H", "Mar", 3, 2), ("K", "May", 5, 4), ("N", "Jul", 7, 6),
        ("U", "Sep", 9, 8), ("Z", "Dec", 12, 11),
    ]
    contracts = []
    for year in [today.year, today.year + 1]:
        for code, label, del_mo, notice_mo in MONTHS:
            last = calendar.monthrange(year, notice_mo)[1]
            d = datetime(year, notice_mo, last).date()
            while d.weekday() >= 5:
                d -= timedelta(days=1)
            yy = str(year)[-2:]
            contracts.append({"code": code, "yf": f"HG{code}{yy}.CMX", "fnd": d})
    contracts.sort(key=lambda c: c["fnd"])

    front = None
    next_mo = None
    for i, c in enumerate(contracts):
        if c["fnd"] >= today:
            front = c
            if i + 1 < len(contracts):
                next_mo = contracts[i + 1]
            break

    if not front:
        return "HG=F", "front"

    # Count trading days to FND
    days = 0
    d = today + timedelta(days=1)
    while d <= front["fnd"]:
        if d.weekday() < 5:
            days += 1
        d += timedelta(days=1)

    # Show next month when <=5 trading days to FND (liquidity migrating)
    if days <= 5 and next_mo:
        return next_mo["yf"], "next"
    return front["yf"], "front"


def _fetch_realtime_price():
    """Get most current COMEX copper price (1-min cache).
    Shows the most liquid contract: front month normally, next month near FND.
    Also tries investing.com first (real-time) before yfinance (5-min delayed).
    Returns dict with price, prev_close, source, and active_contract.
    """
    global _rt_cache
    now = time.time()
    if _rt_cache["price"] and (now - _rt_cache["timestamp"]) < 60:
        return _rt_cache

    # Determine which contract to show based on FND proximity
    active_ticker, active_contract = _get_active_yf_ticker()

    # Method 1: investing.com chart API (real-time, but may be Cloudflare-blocked)
    try:
        import urllib.request
        url = "https://api.investing.com/api/financialdata/8831/historical/chart/?period=P1D&interval=PT5M&pointscount=60"
        req = urllib.request.Request(url, headers={
            "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/131.0.0.0 Safari/537.36",
            "Accept": "application/json",
            "domain-id": "www",
        })
        with urllib.request.urlopen(req, timeout=10) as resp:
            raw = resp.read().decode()
            data = json.loads(raw)

            bars = None
            if isinstance(data, dict) and "data" in data:
                bars = data["data"]
            elif isinstance(data, list):
                bars = data

            if bars and len(bars) > 0:
                latest = bars[-1]
                price = 0
                if isinstance(latest, list) and len(latest) > 4:
                    price = float(latest[4])
                elif isinstance(latest, dict):
                    for key in ("close", "c", "last_close", "last", "price"):
                        if key in latest and latest[key]:
                            price = float(latest[key]); break

                if price > 1:
                    # investing.com tracks May — use May prev settle
                    prev_close = _fetch_prev_settle_yf(active_ticker)
                    if not prev_close:
                        prev_close = _fetch_prev_settle_yf("HG=F")
                    _rt_cache = {"price": round(price, 4), "prev_close": prev_close,
                                 "timestamp": now, "source": "investing.com",
                                 "active_contract": "next"}
                    print(f"[INFO] RT from investing.com: ${price:.4f}")
                    return _rt_cache
    except Exception as e:
        print(f"[WARN] investing.com RT error: {e}")

    # Method 2: yfinance intraday — use the active contract
    try:
        import yfinance as yf
        t = yf.Ticker(active_ticker)
        h = t.history(period="1d", interval="5m")
        if not h.empty:
            h = h.reset_index()
            h.columns = [c if isinstance(c, str) else c[0] for c in h.columns]
            price = float(h.iloc[-1]["Close"])
            if price > 1:
                prev_close = _fetch_prev_settle_yf(active_ticker)
                _rt_cache = {"price": round(price, 4), "prev_close": prev_close,
                             "timestamp": now, "source": "yfinance",
                             "active_contract": active_contract}
                print(f"[INFO] RT from yfinance ({active_ticker}): ${price:.4f} (prev settle: {prev_close})")
                return _rt_cache
    except Exception as e:
        print(f"[WARN] yfinance {active_ticker} intraday error: {e}")

    # Method 3: fallback to HG=F if specific contract failed
    if active_ticker != "HG=F":
        try:
            import yfinance as yf
            t = yf.Ticker("HG=F")
            h = t.history(period="1d", interval="5m")
            if not h.empty:
                h = h.reset_index()
                h.columns = [c if isinstance(c, str) else c[0] for c in h.columns]
                price = float(h.iloc[-1]["Close"])
                if price > 1:
                    prev_close = _fetch_prev_settle_yf("HG=F")
                    _rt_cache = {"price": round(price, 4), "prev_close": prev_close,
                                 "timestamp": now, "source": "yfinance",
                                 "active_contract": "front"}
                    print(f"[INFO] RT fallback from yfinance (HG=F): ${price:.4f}")
                    return _rt_cache
        except Exception as e:
            print(f"[WARN] yfinance HG=F fallback error: {e}")

    return _rt_cache


# ---------------------------------------------------------------------------
# DXY
# ---------------------------------------------------------------------------
_dxy_cache = {"price": None, "change": None, "change_pct": None, "timestamp": 0}

def fetch_dxy():
    global _dxy_cache
    now = time.time()
    if _dxy_cache["price"] and (now - _dxy_cache["timestamp"]) < 300:
        return _dxy_cache
    try:
        import yfinance as yf
        ticker = yf.Ticker("DX-Y.NYB")
        hist = ticker.history(period="5d", interval="1d")
        if not hist.empty:
            hist = hist.reset_index()
            hist.columns = [c if isinstance(c, str) else c[0] for c in hist.columns]
            latest = hist.iloc[-1]; prev = hist.iloc[-2] if len(hist) > 1 else latest
            p = float(latest["Close"]); pc = float(prev["Close"]); ch = p - pc
            _dxy_cache = {"price": round(p, 2), "change": round(ch, 2),
                          "change_pct": round((ch / pc) * 100, 2) if pc else 0, "timestamp": now}
    except Exception as e:
        print(f"[WARN] DXY fetch error: {e}")
    return _dxy_cache


# ---------------------------------------------------------------------------
# FRED API — Fed Funds target rate (24h cache)
# ---------------------------------------------------------------------------
_fred_cache = {"upper": None, "lower": None, "midpoint": None, "rate_str": None, "timestamp": 0}

def fetch_fred_fed_rate():
    """Fetch current Fed Funds target rate from FRED (DFEDTARU/DFEDTARL). 24h cache."""
    global _fred_cache
    now = time.time()
    if _fred_cache["upper"] is not None and (now - _fred_cache["timestamp"]) < 86400:
        return _fred_cache

    api_key = CFG["FRED_API_KEY"]
    if not api_key:
        return _fred_cache

    import urllib.request
    upper = None; lower = None
    for series_id in ("DFEDTARU", "DFEDTARL"):
        try:
            url = (f"https://api.stlouisfed.org/fred/series/observations"
                   f"?series_id={series_id}&api_key={api_key}"
                   f"&file_type=json&sort_order=desc&limit=5")
            req = urllib.request.Request(url, headers={"User-Agent": "GeometDashboard/1.0"})
            with urllib.request.urlopen(req, timeout=10) as resp:
                data = json.loads(resp.read().decode())
                for obs in data.get("observations", []):
                    if obs.get("value") and obs["value"] != ".":
                        val = float(obs["value"])
                        if series_id == "DFEDTARU":
                            upper = val
                        else:
                            lower = val
                        break
        except Exception as e:
            print(f"[WARN] FRED {series_id} error: {e}")

    if upper is not None and lower is not None:
        midpoint = round((upper + lower) / 2, 3)
        rate_str = f"{lower:.2f}-{upper:.2f}"
        _fred_cache = {"upper": upper, "lower": lower, "midpoint": midpoint,
                       "rate_str": rate_str, "timestamp": now}
        print(f"[INFO] FRED fed funds: {rate_str}% (midpoint {midpoint}%)")
    return _fred_cache


# ---------------------------------------------------------------------------
# FED FUNDS / RATE EXPECTATIONS
# ---------------------------------------------------------------------------
_fed_cache = {"data": None, "timestamp": 0}

def fetch_fed_data():
    global _fed_cache
    now = time.time()
    if _fed_cache["data"] and (now - _fed_cache["timestamp"]) < 900:
        return _fed_cache["data"]

    # Use FRED data if available, otherwise fall back to static config
    fred = fetch_fred_fed_rate()
    if fred.get("rate_str"):
        current_rate = fred["rate_str"]
        midpoint = fred["midpoint"]
    else:
        current_rate = CFG["FED_FUNDS_RATE"]
        midpoint = CFG["FED_FUNDS_MIDPOINT"]

    result = {
        "current_rate": current_rate,
        "midpoint": midpoint,
    }
    try:
        import yfinance as yf
        # Try fed funds futures for upcoming months
        # ZQ contracts: price = 100 - implied rate
        implied_rates = {}
        month_labels = {
            "ZQH26.CBT": "Mar 26", "ZQJ26.CBT": "Apr 26", "ZQK26.CBT": "May 26",
            "ZQM26.CBT": "Jun 26", "ZQN26.CBT": "Jul 26", "ZQQ26.CBT": "Aug 26",
            "ZQU26.CBT": "Sep 26", "ZQV26.CBT": "Oct 26", "ZQX26.CBT": "Nov 26",
            "ZQZ26.CBT": "Dec 26",
        }
        found_any = False
        for sym, label in month_labels.items():
            try:
                t = yf.Ticker(sym)
                h = t.history(period="5d")
                if not h.empty:
                    price = float(h.iloc[-1]["Close"])
                    implied = round(100 - price, 3)
                    implied_rates[label] = implied
                    found_any = True
            except: continue

        if found_any:
            result["implied_rates"] = implied_rates
            # Find first month where rate drops by 25bp+
            first_cut = None; total_cuts = 0
            for label in ["Mar 26","Apr 26","May 26","Jun 26","Jul 26","Aug 26","Sep 26","Oct 26","Nov 26","Dec 26"]:
                if label in implied_rates:
                    cuts = max(0, round((midpoint - implied_rates[label]) / 0.25))
                    if cuts > 0 and not first_cut:
                        first_cut = label
                    total_cuts = max(total_cuts, cuts)
            result["first_cut"] = first_cut if first_cut else "None priced"
            result["total_cuts_2026"] = total_cuts
        else:
            # Fallback: 10Y yield as macro context
            try:
                t10 = yf.Ticker("^TNX")
                h10 = t10.history(period="5d")
                if not h10.empty:
                    h10 = h10.reset_index()
                    h10.columns = [c if isinstance(c, str) else c[0] for c in h10.columns]
                    y10 = float(h10.iloc[-1]["Close"])
                    prev10 = float(h10.iloc[-2]["Close"]) if len(h10) > 1 else y10
                    result["yield_10y"] = round(y10, 2)
                    result["yield_10y_change"] = round(y10 - prev10, 2)
            except: pass
    except Exception as e:
        print(f"[WARN] Fed data error: {e}")

    _fed_cache = {"data": result, "timestamp": now}
    return result


# ---------------------------------------------------------------------------
# CHINA / SHFE STATUS
# ---------------------------------------------------------------------------
def get_china_status():
    now = datetime.now()
    lny_start = datetime(2026, 2, 15)
    lny_end = datetime(2026, 2, 23, 23, 59, 59)
    if lny_start <= now <= lny_end:
        days_left = (lny_end - now).days
        return {"status": "CLOSED", "reason": "Lunar New Year",
                "detail": f"SHFE closed \u2014 returns Feb 24 ({days_left}d)", "color": "red", "thin_liquidity": True}
    wd = now.weekday(); hr = now.hour
    if wd >= 5:
        return {"status": "CLOSED", "reason": "Weekend", "detail": "SHFE closed \u2014 weekend",
                "color": "yellow", "thin_liquidity": wd == 6 and hr >= 17}
    if hr >= 19 or hr < 2:
        return {"status": "OPEN", "reason": "Night session", "detail": "SHFE night session active",
                "color": "green", "thin_liquidity": False}
    if 7 <= hr <= 14:
        return {"status": "CLOSED", "reason": "Between sessions", "detail": "SHFE between sessions",
                "color": "yellow", "thin_liquidity": False}
    return {"status": "CLOSED", "reason": "Off hours", "detail": "SHFE closed",
            "color": "yellow", "thin_liquidity": False}


# ---------------------------------------------------------------------------
# LME STATUS — electronic + ring hours (London time)
# ---------------------------------------------------------------------------
def get_lme_status():
    """LME market hours status based on London time.
    LME Select (electronic): 01:00-19:00 London
    Official Ring session: 11:40-17:00 London
    """
    from datetime import timezone
    import zoneinfo
    try:
        london = zoneinfo.ZoneInfo("Europe/London")
    except Exception:
        # Fallback: UTC offset approximation (GMT/BST)
        london = timezone.utc
    now_london = datetime.now(london)
    wd = now_london.weekday()
    hr = now_london.hour
    mn = now_london.minute
    t = hr * 60 + mn  # minutes since midnight

    if wd >= 5:
        return {"status": "CLOSED", "detail": "LME CLOSED", "color": "yellow", "session": "weekend"}

    # Ring session: 11:40 (700) - 17:00 (1020) London
    if 700 <= t < 1020:
        return {"status": "RING", "detail": "LME RING", "color": "green", "session": "ring"}

    # LME Select electronic: 01:00 (60) - 19:00 (1140) London
    if 60 <= t < 1140:
        return {"status": "OPEN", "detail": "LME OPEN", "color": "green", "session": "electronic"}

    return {"status": "CLOSED", "detail": "LME CLOSED", "color": "yellow", "session": "closed"}


# ---------------------------------------------------------------------------
# SPREAD HISTORY
# ---------------------------------------------------------------------------
def load_spread_history():
    if SPREAD_HISTORY.exists():
        try:
            with open(SPREAD_HISTORY) as f: return json.load(f)
        except: pass
    return []

def save_spread_entry(comex, lme, spread):
    history = load_spread_history()
    today = datetime.now().strftime("%Y-%m-%d")
    if history and history[-1].get("date") == today:
        history[-1] = {"date": today, "comex": comex, "lme": lme, "spread": spread}
    else:
        history.append({"date": today, "comex": comex, "lme": lme, "spread": spread})
    history = history[-180:]
    try:
        with open(SPREAD_HISTORY, "w") as f: json.dump(history, f)
    except: pass
    return history

def compute_spread_intelligence(history, current_spread):
    if not history or current_spread is None: return None
    spreads = [h["spread"] for h in history if h.get("spread") is not None]
    if len(spreads) < 3:
        return {"history_days": len(spreads)}
    s30 = spreads[-30:] if len(spreads) >= 30 else spreads
    below30 = sum(1 for s in s30 if s < current_spread)
    pct30 = round(below30 / len(s30) * 100, 1)
    streak = 0; direction = None
    for i in range(len(spreads)-1, 0, -1):
        diff = spreads[i] - spreads[i-1]
        if diff > 0.001:
            if direction == "widening" or direction is None: streak += 1; direction = "widening"
            else: break
        elif diff < -0.001:
            if direction == "narrowing" or direction is None: streak += 1; direction = "narrowing"
            else: break
        else: break
    return {"pct_30d": pct30, "streak": streak, "streak_direction": direction,
            "range_30d_min": round(min(s30), 4), "range_30d_max": round(max(s30), 4), "history_days": len(spreads)}


# ---------------------------------------------------------------------------
# AUTO SUPPORT/RESISTANCE
# ---------------------------------------------------------------------------
def calc_support_resistance(closes, highs, lows):
    if len(closes) < 20: return {"support": [], "resistance": []}
    current = closes[-1]; swing_highs = []; swing_lows = []; lookback = 5
    for i in range(lookback, len(highs) - lookback):
        if highs[i] == max(highs[i-lookback:i+lookback+1]): swing_highs.append(round(highs[i], 4))
        if lows[i] == min(lows[i-lookback:i+lookback+1]): swing_lows.append(round(lows[i], 4))

    def cluster(levels, threshold=0.03):
        if not levels: return []
        levels = sorted(levels); clusters = []; cc = [levels[0]]
        for i in range(1, len(levels)):
            if levels[i] - cc[-1] < threshold: cc.append(levels[i])
            else: clusters.append(round(sum(cc)/len(cc), 4)); cc = [levels[i]]
        clusters.append(round(sum(cc)/len(cc), 4)); return clusters

    support = cluster([s for s in swing_lows if s < current])
    resistance = cluster([r for r in swing_highs if r > current])
    support = sorted(support, reverse=True)[:3]
    resistance = sorted(resistance)[:3]
    return {
        "support": [{"level": s, "distance": round(current-s, 4), "distance_pct": round((current-s)/current*100, 2)} for s in support],
        "resistance": [{"level": r, "distance": round(r-current, 4), "distance_pct": round((r-current)/current*100, 2)} for r in resistance],
    }


# ---------------------------------------------------------------------------
# FIX WINDOW SCORING
# ---------------------------------------------------------------------------
def calc_fix_window(md, sig):
    """Composite score: should you fix/price against unpriced longs right now?"""
    if not md or not sig: return None
    score = 50  # neutral baseline

    # 1. Percentile position (higher = better for fixing)
    pct90 = md.get("pct_90d", 50)
    if pct90 >= 85: score += 18
    elif pct90 >= 70: score += 12
    elif pct90 >= 55: score += 5
    elif pct90 <= 15: score -= 18
    elif pct90 <= 30: score -= 12
    elif pct90 <= 45: score -= 5

    # 2. Momentum (rising = better for fixing)
    roc = md.get("roc", {})
    r5 = roc.get("5d", {}).get("pct", 0)
    if r5 > 3: score += 12
    elif r5 > 1: score += 6
    elif r5 < -3: score -= 12
    elif r5 < -1: score -= 6

    # 3. DXY (dollar weak = bullish copper = better for fixing)
    dxy = md.get("dxy", {})
    dch = dxy.get("change_pct", 0)
    if dch < -0.5: score += 10
    elif dch < -0.2: score += 5
    elif dch > 0.5: score -= 10
    elif dch > 0.2: score -= 5

    # 4. Trend
    trend = sig.get("trend", "")
    ts = sig.get("trend_strength", "")
    if trend == "UPTREND" and ts == "strong": score += 10
    elif trend == "UPTREND": score += 5
    elif trend == "DOWNTREND" and ts == "strong": score -= 10
    elif trend == "DOWNTREND": score -= 5

    # 5. Price vs fix target
    ft = CFG["FIX_TARGET"]; p = md["price"]
    if p >= ft + 0.10: score += 15
    elif p >= ft: score += 12
    elif p >= ft - 0.05: score += 5
    elif p < ft - 0.20: score -= 5

    # 6. Spread (COMEX premium = favorable for COMEX fixing)
    spread = md.get("comex_lme_spread")
    if spread is not None:
        if spread > 0.10: score += 5
        elif spread < -0.10: score -= 3

    # 7. China status (thin liquidity = less reliable signals)
    china = md.get("china", {})
    if china.get("thin_liquidity"): score -= 5

    score = max(0, min(100, score))

    if score >= 80: label, color = "STRONG FIX", "green"
    elif score >= 65: label, color = "FAVORABLE", "green"
    elif score >= 45: label, color = "NEUTRAL", "yellow"
    elif score >= 25: label, color = "UNFAVORABLE", "orange"
    else: label, color = "HOLD / BUY", "red"

    # Build factors list
    factors = []
    if pct90 >= 75: factors.append(f"Near 90d highs ({pct90}th pctl)")
    elif pct90 <= 25: factors.append(f"Near 90d lows ({pct90}th pctl)")
    if r5 > 1: factors.append(f"Momentum up ({r5:+.1f}%)")
    elif r5 < -1: factors.append(f"Momentum down ({r5:+.1f}%)")
    if dch < -0.3: factors.append("Dollar weakening")
    elif dch > 0.3: factors.append("Dollar strengthening")
    if p >= ft: factors.append(f"Above ${ft:.2f} target")
    elif p >= ft - 0.10: factors.append(f"Near ${ft:.2f} target")
    if china.get("thin_liquidity"): factors.append("Thin liquidity")
    if trend == "UPTREND": factors.append("Uptrend")
    elif trend == "DOWNTREND": factors.append("Downtrend")

    return {"score": score, "label": label, "color": color, "factors": factors}


def calc_fixable_orders(pos, md):
    """Determine which unpriced orders are fixable now based on exchange hours."""
    if not pos or not md:
        return None
    shipped = pos.get("sales_unpriced_shipped", [])
    unshipped = pos.get("sales_unpriced_unshipped", [])
    all_unpriced = shipped + unshipped
    if not all_unpriced:
        return None

    lme = md.get("lme", {})
    lme_open = lme.get("status") in ("OPEN", "RING") if lme else False
    # COMEX electronic is essentially 23h/day Sun-Fri, assume open if market data exists
    comex_open = md.get("price") is not None

    fixable = 0; fixable_lbs = 0; blocked_lme = 0; blocked_lme_lbs = 0
    shipped_fixable = []
    for sale in all_unpriced:
        basis = sale.get("basis", "COMEX")
        lbs = sale.get("open_lbs", sale.get("lbs", 0))
        if basis == "LME" and not lme_open:
            blocked_lme += 1; blocked_lme_lbs += lbs
        else:
            fixable += 1; fixable_lbs += lbs
            if sale in shipped:
                shipped_fixable.append(sale)

    return {
        "fixable_count": fixable, "fixable_lbs": round(fixable_lbs),
        "blocked_lme_count": blocked_lme, "blocked_lme_lbs": round(blocked_lme_lbs),
        "total_unpriced": len(all_unpriced),
        "shipped_fixable": shipped_fixable,
    }


# ---------------------------------------------------------------------------
# WAREHOUSE STOCKS — CME scraper with 24h cache
# ---------------------------------------------------------------------------
SHORT_TON_TO_MT = 0.907185
_cme_wh_cache = {"data": None, "timestamp": 0}

def fetch_cme_warehouse():
    """Download and parse CME Copper_Stocks.xls for live warehouse data."""
    global _cme_wh_cache
    now = time.time()
    if _cme_wh_cache["data"] and (now - _cme_wh_cache["timestamp"]) < 86400:
        return _cme_wh_cache["data"]

    try:
        import urllib.request, tempfile, xlrd
        url = "https://www.cmegroup.com/delivery_reports/Copper_Stocks.xls"
        req = urllib.request.Request(url, headers={"User-Agent": "GeometDashboard/1.0"})
        with urllib.request.urlopen(req, timeout=15) as resp:
            xls_data = resp.read()

        tmp = tempfile.NamedTemporaryFile(suffix=".xls", delete=False)
        tmp.write(xls_data); tmp.close()

        wb = xlrd.open_workbook(tmp.name)
        ws = wb.sheet_by_index(0)

        # Extract activity date from row 8 (e.g. "Activity Date: 2/18/2026")
        activity_date = ""
        for r in range(min(10, ws.nrows)):
            val = str(ws.cell_value(r, 6)).strip()
            if "Activity Date" in val:
                activity_date = val.replace("Activity Date:", "").strip()
                break

        # Find TOTAL COPPER row for totals
        total_today_st = 0; prev_total_st = 0
        for r in range(ws.nrows):
            label = str(ws.cell_value(r, 0)).strip()
            if label == "TOTAL COPPER":
                prev_total_st = float(ws.cell_value(r, 2)) if ws.cell_value(r, 2) else 0
                total_today_st = float(ws.cell_value(r, 7)) if ws.cell_value(r, 7) else 0
                break

        os.unlink(tmp.name)

        if total_today_st <= 0:
            return None

        total_mt = int(round(total_today_st * SHORT_TON_TO_MT))
        prev_mt = int(round(prev_total_st * SHORT_TON_TO_MT))
        net_change_mt = total_mt - prev_mt
        if net_change_mt > 0:
            trend = "building"
        elif net_change_mt < 0:
            trend = "drawing"
        else:
            trend = "stable"

        result = {
            "mt": total_mt, "lbs": int(total_mt * MT_TO_LB),
            "date": activity_date, "trend": trend,
            "short_tons": int(total_today_st),
            "net_change_mt": net_change_mt,
            "source": "cme",
        }
        _cme_wh_cache = {"data": result, "timestamp": now}
        print(f"[INFO] CME warehouse: {total_mt:,} MT ({trend}, {activity_date})")
        return result
    except Exception as e:
        print(f"[WARN] CME warehouse scrape error: {e}")
        return None

def get_warehouse_data():
    # Try live CME data first, fall back to static config
    live = fetch_cme_warehouse()
    if live:
        return live
    wh = CFG.get("COMEX_WAREHOUSE_MT", 0)
    if not wh: return None
    return {
        "mt": wh, "lbs": int(wh * MT_TO_LB),
        "date": CFG.get("COMEX_WAREHOUSE_DATE", ""),
        "trend": CFG.get("COMEX_WAREHOUSE_TREND", ""),
        "source": "config",
    }


# ---------------------------------------------------------------------------
# CONTRACT ROLL + OPEN INTEREST
# ---------------------------------------------------------------------------
_roll_cache = {"data": None, "timestamp": 0}
OI_HISTORY = DATA_DIR / "oi_history.json"

def get_contract_roll(copper_price=None):
    """COMEX copper contract roll status, calendar spread, and open interest."""
    global _roll_cache
    now = time.time()
    if _roll_cache["data"] and (now - _roll_cache["timestamp"]) < 300:
        cached = _roll_cache["data"].copy()
        if copper_price is not None:
            cached["front_price"] = round(copper_price, 4)
            if cached.get("next_price"):
                spread = round(cached["next_price"] - copper_price, 4)
                cached["calendar_spread"] = spread
                cached["market_structure"] = "contango" if spread > 0.001 else "backwardation" if spread < -0.001 else "flat"
        return cached

    today = datetime.now().date()

    # COMEX copper active months: H=Mar, K=May, N=Jul, U=Sep, Z=Dec
    MONTHS = [
        ("H", "Mar", 3, 2),
        ("K", "May", 5, 4),
        ("N", "Jul", 7, 6),
        ("U", "Sep", 9, 8),
        ("Z", "Dec", 12, 11),
    ]

    def last_biz_day(year, month):
        last = calendar.monthrange(year, month)[1]
        d = datetime(year, month, last).date()
        while d.weekday() >= 5:
            d -= timedelta(days=1)
        return d

    def trading_days_until(target):
        if today >= target:
            return 0
        count = 0
        d = today + timedelta(days=1)
        while d <= target:
            if d.weekday() < 5:
                count += 1
            d += timedelta(days=1)
        return count

    contracts = []
    for year in [today.year, today.year + 1]:
        for code, label, del_mo, notice_mo in MONTHS:
            fnd = last_biz_day(year, notice_mo)
            yy = str(year)[-2:]
            contracts.append({
                "code": code, "label": f"{label} {yy}",
                "ticker": f"HG{code}{yy}",
                "yf_ticker": f"HG{code}{yy}.CMX",
                "fnd": fnd, "fnd_str": fnd.strftime("%b %d"),
                "year": year,
            })
    contracts.sort(key=lambda c: c["fnd"])

    front = None; next_mo = None
    for i, c in enumerate(contracts):
        if c["fnd"] >= today:
            front = c
            if i + 1 < len(contracts):
                next_mo = contracts[i + 1]
            break
    if not front:
        return None

    days_to_fnd = trading_days_until(front["fnd"])

    if today == front["fnd"]:
        roll_status = "FIRST NOTICE DAY"
        roll_urgency = "critical"
        roll_color = "red"
    elif days_to_fnd <= 2:
        roll_status = f"{days_to_fnd} trading day{'s' if days_to_fnd != 1 else ''} to FND"
        roll_urgency = "critical"
        roll_color = "red"
    elif days_to_fnd <= 5:
        roll_status = f"{days_to_fnd} trading days to FND"
        roll_urgency = "warning"
        roll_color = "orange"
    elif days_to_fnd <= 10:
        roll_status = f"{days_to_fnd} trading days to FND"
        roll_urgency = "attention"
        roll_color = "yellow"
    else:
        roll_status = f"{days_to_fnd} trading days to FND"
        roll_urgency = "normal"
        roll_color = "green"

    result = {
        "front_month": {"label": front["label"], "ticker": front["ticker"], "fnd": front["fnd_str"]},
        "days_to_fnd": days_to_fnd,
        "roll_status": roll_status,
        "roll_urgency": roll_urgency,
        "roll_color": roll_color,
    }
    if next_mo:
        result["next_month"] = {"label": next_mo["label"], "ticker": next_mo["ticker"], "fnd": next_mo["fnd_str"]}
    if copper_price is not None:
        result["front_price"] = round(copper_price, 4)
    else:
        # Fetch front month price from yfinance (needed when RT source is next month)
        try:
            import yfinance as yf
            t = yf.Ticker(front["yf_ticker"])
            h = t.history(period="1d", interval="5m")
            if h.empty:
                h = t.history(period="5d")
            if not h.empty:
                h = h.reset_index()
                h.columns = [c if isinstance(c, str) else c[0] for c in h.columns]
                result["front_price"] = round(float(h.iloc[-1]["Close"]), 4)
                print(f"[INFO] Front month {front['ticker']} (yf): ${result['front_price']:.4f}")
        except Exception as e:
            print(f"[WARN] Front month price error: {e}")

    # Fetch next month contract price for calendar spread
    if next_mo:
        try:
            import yfinance as yf
            t = yf.Ticker(next_mo["yf_ticker"])
            # Try intraday first for most current price
            h = t.history(period="1d", interval="5m")
            if h.empty:
                h = t.history(period="5d")  # fallback to daily
            if not h.empty:
                h = h.reset_index()
                h.columns = [c if isinstance(c, str) else c[0] for c in h.columns]
                result["next_price"] = round(float(h.iloc[-1]["Close"]), 4)
                print(f"[INFO] Next month {next_mo['ticker']}: ${result['next_price']:.4f}")
        except Exception as e:
            print(f"[WARN] Next month price error: {e}")

    # Calendar spread
    if copper_price and result.get("next_price"):
        spread = round(result["next_price"] - copper_price, 4)
        result["calendar_spread"] = spread
        result["market_structure"] = "contango" if spread > 0.001 else "backwardation" if spread < -0.001 else "flat"

    # Open interest
    try:
        import yfinance as yf
        t = yf.Ticker("HG=F")
        info = t.info or {}
        oi = info.get("openInterest")
        if oi and oi > 0:
            oi_history = _save_oi(oi)
            oi_data = {"total": oi, "source": "yfinance"}
            trend = _compute_oi_trend(oi_history, oi)
            if trend:
                oi_data.update(trend)
            result["open_interest"] = oi_data
            print(f"[INFO] Open interest: {oi:,} contracts")
    except Exception as e:
        print(f"[WARN] Open interest error: {e}")

    _roll_cache = {"data": result, "timestamp": now}
    return result


def _save_oi(oi):
    try:
        history = []
        if OI_HISTORY.exists():
            with open(OI_HISTORY) as f:
                history = json.load(f)
        today_str = datetime.now().strftime("%Y-%m-%d")
        if history and history[-1].get("date") == today_str:
            history[-1] = {"date": today_str, "oi": oi}
        else:
            history.append({"date": today_str, "oi": oi})
        history = history[-30:]
        with open(OI_HISTORY, "w") as f:
            json.dump(history, f)
        return history
    except:
        return []


def _compute_oi_trend(history, current_oi):
    if not history or len(history) < 2:
        return None
    recent = history[-5:] if len(history) >= 5 else history
    first_oi = recent[0]["oi"]
    if first_oi <= 0:
        return None
    change = current_oi - first_oi
    change_pct = round((change / first_oi) * 100, 1)
    trend = "building" if change > 0 else "declining" if change < 0 else "stable"
    return {"trend": trend, "change_5d": change, "change_5d_pct": change_pct, "history_days": len(history)}


# ---------------------------------------------------------------------------
# HEDGE SPREADSHEET READER
# ---------------------------------------------------------------------------
def is_real_file(filepath):
    """Check if file is actually downloaded (not a OneDrive placeholder)."""
    try:
        with open(filepath, 'rb') as f:
            header = f.read(4)
        return header == b'PK\x03\x04'  # Valid xlsx/zip header
    except:
        return False

def trigger_onedrive_download(filepath):
    """Ask OneDrive to download a cloud-only placeholder file."""
    import subprocess
    try:
        subprocess.run(["brctl", "download", filepath], timeout=5, capture_output=True)
        print(f"[INFO] Triggered OneDrive download: {os.path.basename(filepath)}")
    except Exception as e:
        print(f"[WARN] brctl download failed: {e}")

def find_latest_hedge_file():
    # Scan OneDrive sync folder + local data dir
    onedrive_dirs = [
        os.path.expanduser(
            "~/Library/CloudStorage/OneDrive-GeometRecycle/"
            "Pricing_Hedge - Hedge Worksheet"
        ),
        # Legacy path (older OneDrive versions)
        os.path.expanduser(
            "~/Library/Group Containers/UBF8T346G9.OneDriveSyncClientSuite/"
            "OneDrive - Geomet Recycle.noindex/OneDrive - Geomet Recycle/"
            "Pricing_Hedge - Hedge Worksheet"
        ),
    ]
    files = glob.glob(str(DATA_DIR / "Hedge*.xlsx"))
    for onedrive_dir in onedrive_dirs:
        try:
            files += glob.glob(os.path.join(onedrive_dir, "Hedge*.xlsx"))
        except PermissionError:
            print(f"[WARN] Permission denied: {onedrive_dir} — grant Full Disk Access to Python")
    if not files: return None
    def extract_date(f):
        base = os.path.basename(f)
        match = re.search(r'Hedge(\d{8})', base)
        if match:
            try: return datetime.strptime(match.group(1), "%m%d%Y")
            except: pass
        return datetime.fromtimestamp(os.path.getmtime(f))
    files.sort(key=extract_date, reverse=True)

    # Try latest file first — trigger download if placeholder
    for filepath in files:
        if is_real_file(filepath):
            return filepath
        # It's a placeholder — try to trigger OneDrive download
        trigger_onedrive_download(filepath)
        # Wait up to 15 seconds for download
        for i in range(15):
            time.sleep(1)
            if is_real_file(filepath):
                print(f"[INFO] OneDrive download complete: {os.path.basename(filepath)}")
                return filepath
        print(f"[WARN] Skipping placeholder: {os.path.basename(filepath)}")
    return None

def read_hedge_spreadsheet(filepath):
    try:
        import openpyxl
        wb = openpyxl.load_workbook(filepath, data_only=True)
        result = {
            "net_lbs": 0, "avg_cost": 0, "hedge_lbs": 0,
            "priced_sales_lbs": 0, "priced_sales_avg": 0, "unpriced_sales_lbs": 0,
            "total_inv_po": 0, "comex_futures": 0, "lme_futures": 0,
            "updated": "", "source_file": os.path.basename(filepath),
            "sales_priced_unshipped": [], "sales_unpriced_shipped": [], "sales_unpriced_unshipped": [],
            "sales_priced_unshipped_lbs": 0, "sales_unpriced_shipped_lbs": 0, "sales_unpriced_unshipped_lbs": 0,
            "inv_by_commodity": {"BB": 0, "#1": 0, "#2": 0, "Chops": 0},
            "sales_by_commodity": {"BB": 0, "#1": 0, "#2": 0, "Chops": 0},
        }

        if "POSITION" in wb.sheetnames:
            ws = wb["POSITION"]
            grid = [[c if c is not None else "" for c in row] for row in ws.iter_rows(values_only=True)]
            for i, row in enumerate(grid):
                for j, cell in enumerate(row):
                    s = str(cell).strip()
                    if i == 0 and j == 1 and isinstance(cell, (int, float)): result["net_lbs"] = float(cell)
                    # avg_cost calculated below from per-commodity costs
                    if s == "COMEX FUTURES":
                        try:
                            v = row[1] if j == 0 else (grid[i][j+1] if j+1<len(row) else 0)
                            if isinstance(v, (int, float)): result["comex_futures"] = float(v)
                        except: pass
                    if s == "LME FUTURES":
                        try:
                            v = row[1] if j == 0 else (grid[i][j+1] if j+1<len(row) else 0)
                            if isinstance(v, (int, float)): result["lme_futures"] = float(v)
                        except: pass
                    if s == "PRICED SO OS" and j == 0:
                        try: result["priced_sales_lbs"] = float(row[1]) if row[1] else 0
                        except: pass
            # Per-commodity avg costs from Power BI (until Jorge adds cost column to spreadsheet)
            _COST_PER_LB = {
                "CU1": 5.280171, "CU2": 5.134602, "CU2DIRTY": 5.134602, "CUBB": 5.455980,
                "CAT5": 2.024765, "CUINS1": 3.041050, "CUINS2": 2.093361,
                "MCM": 4.199354, "THHN": 3.914800, "WAVEOPENCU": 2.355245,
                "CUCHOP CUBB": 4.616107, "CUCHOP1A_M": 4.616107, "CUCHOPS2": 4.616107,
            }
            total_cost = 0; total_cu_lbs = 0
            # Extract per-commodity inventory from solid inventory columns (col 2=item, col 3=weight, col 5=CuUnits)
            for row in grid[7:25]:
                if len(row) > 5 and isinstance(row[2], str) and isinstance(row[5], (int, float)):
                    item = row[2].strip().upper()
                    cu = float(row[5])
                    wt = float(row[3]) if isinstance(row[3], (int, float)) else cu
                    if item in _COST_PER_LB:
                        total_cost += wt * _COST_PER_LB[item]; total_cu_lbs += cu
                    if item.startswith("CUBB") and "CHOP" not in item:
                        result["inv_by_commodity"]["BB"] += cu
                    elif item.startswith("CU1"):
                        result["inv_by_commodity"]["#1"] += cu
                    elif item.startswith("CU2"):
                        result["inv_by_commodity"]["#2"] += cu
                    elif item.startswith("CUCHOP"):
                        result["inv_by_commodity"]["Chops"] += cu
            # Add ICW inventory (at projected recovery) to Chops — insulated wire becomes chops when processed
            # Also include ICW in weighted avg cost
            for i, row in enumerate(grid):
                for j, cell in enumerate(row):
                    if str(cell).strip() == "ICW INV" and j + 1 < len(row) and isinstance(row[j + 1], (int, float)):
                        icw_cu = float(row[j + 1])
                        result["inv_by_commodity"]["Chops"] += icw_cu
                        result["icw_cu_lbs"] = icw_cu
                        result["chops_solid_lbs"] = result["inv_by_commodity"]["Chops"] - icw_cu
            # ICW per-item costs (col 9=item, col 10=weight, col 12=CuUnits)
            for row in grid[7:25]:
                if len(row) > 12 and isinstance(row[9], str) and isinstance(row[10], (int, float)):
                    item = row[9].strip().upper()
                    wt = float(row[10])
                    cu = float(row[12]) if isinstance(row[12], (int, float)) else 0
                    if item in _COST_PER_LB:
                        total_cost += wt * _COST_PER_LB[item]; total_cu_lbs += cu
            # Avg cost per lb of recovered copper (total $ paid / total Cu lbs out)
            result["avg_cost"] = round(total_cost / total_cu_lbs, 6) if total_cu_lbs > 0 else 0

        shipped_orders = {}
        if "O SOLID SALE" in wb.sheetnames:
            ws = wb["O SOLID SALE"]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or not row[0]: continue
                try:
                    so_num = str(row[2]).strip() if row[2] else ""
                    shipped_lbs = float(row[8]) if row[8] else 0
                    if so_num and shipped_lbs > 0:
                        shipped_orders[so_num] = shipped_orders.get(so_num, 0) + shipped_lbs
                except: continue

        if "SOSOLIDS" in wb.sheetnames:
            ws = wb["SOSOLIDS"]
            priced_total = 0; priced_value = 0; unpriced_total = 0
            for row in ws.iter_rows(min_row=3, values_only=True):
                if not row or not row[0]: continue
                try:
                    order = str(row[1]).strip() if row[1] else ""
                    consumer = str(row[2]) if row[2] else ""
                    poref = str(row[3]) if row[3] else ""
                    option = str(row[4]) if row[4] else ""
                    commodity = str(row[5]) if row[5] else ""
                    total_tons = float(row[6]) if row[6] else 0
                    open_priced = float(row[7]) if row[7] else 0
                    priced_tons = float(row[8]) if row[8] else 0
                    final_price = float(row[15]) if row[15] else 0
                    fix_month = str(row[12]) if row[12] else ""
                    spread = float(row[14]) if row[14] else 0
                    basis = "COMEX"
                    if "LME" in poref.upper() or "LME" in option.upper() or "LME" in fix_month.upper():
                        basis = "LME"
                    sale = {"order": order, "consumer": consumer, "commodity": commodity,
                            "lbs": total_tons, "option": option, "basis": basis, "spread": spread, "poref": poref}
                    if priced_tons > 0:
                        priced_total += priced_tons; priced_value += priced_tons * final_price
                        sale["price"] = round(final_price, 4); sale["status"] = "PRICED"
                        sale["priced_lbs"] = priced_tons
                        if order not in shipped_orders:
                            result["sales_priced_unshipped"].append(sale)
                            result["sales_priced_unshipped_lbs"] += priced_tons
                    elif open_priced > 0:
                        unpriced_total += open_priced; sale["status"] = "UNPRICED"
                        sale["open_lbs"] = open_priced
                        if order in shipped_orders:
                            sale["shipped_lbs"] = shipped_orders[order]
                            result["sales_unpriced_shipped"].append(sale)
                            result["sales_unpriced_shipped_lbs"] += open_priced
                        else:
                            result["sales_unpriced_unshipped"].append(sale)
                            result["sales_unpriced_unshipped_lbs"] += open_priced
                    sale_lbs = priced_tons + open_priced
                    cu = commodity.upper()
                    if cu.startswith("CUBB") and "CHOP" not in cu:
                        result["sales_by_commodity"]["BB"] += sale_lbs
                    elif cu.startswith("CU1"):
                        result["sales_by_commodity"]["#1"] += sale_lbs
                    elif cu.startswith("CU2"):
                        result["sales_by_commodity"]["#2"] += sale_lbs
                    elif "CHOP" in cu:
                        result["sales_by_commodity"]["Chops"] += sale_lbs
                except (TypeError, ValueError, IndexError): continue
            if priced_total > 0: result["priced_sales_avg"] = round(priced_value / priced_total, 4)
            result["priced_sales_lbs"] = priced_total; result["unpriced_sales_lbs"] = unpriced_total

        if "Report" in wb.sheetnames:
            ws = wb["Report"]
            for row in ws.iter_rows(values_only=True):
                if row and row[0]:
                    label = str(row[0]).strip(); val = row[1] if len(row) > 1 else None
                    if label == "Total Inv/PO" and isinstance(val, (int, float)):
                        result["total_inv_po"] = float(val)
                    if label == "Inventory Copper" and isinstance(val, (int, float)):
                        result["inventory_cu_lbs"] = float(val)
                    if label.startswith("PO Waiting") and isinstance(val, (int, float)):
                        result["po_lbs"] = float(val)

        result["hedge_lbs"] = abs(result["comex_futures"]) + abs(result["lme_futures"])
        match = re.search(r'Hedge(\d{8})', os.path.basename(filepath))
        if match:
            try: result["updated"] = datetime.strptime(match.group(1), "%m%d%Y").strftime("%Y-%m-%d")
            except: result["updated"] = match.group(1)
        wb.close()
        return result
    except Exception as e:
        print(f"[ERROR] read_hedge: {e}")
        import traceback; traceback.print_exc()
        return None


# ---------------------------------------------------------------------------
# COMEX MARKET DATA + PRICE CONTEXT
# ---------------------------------------------------------------------------
_copper_cache = {"data": None, "timestamp": 0}
COPPER_CACHE_TTL = 300  # 5 minutes

def _fetch_ohlc_investiny():
    """Fetch ~1 year COMEX copper OHLC from investing.com via investiny."""
    from investiny import historical_data
    now = datetime.now()
    one_year_ago = now - timedelta(days=365)
    data = historical_data(
        investing_id=8831,
        from_date=one_year_ago.strftime("%m/%d/%Y"),
        to_date=now.strftime("%m/%d/%Y"),
    )
    if not data or not data.get("close") or len(data["close"]) < 20:
        return None
    n = len(data["close"])
    dates = [datetime.strptime(d, "%m/%d/%Y") for d in data["date"]]
    closes = [float(c) for c in data["close"]]
    highs = [float(h) for h in data["high"]]
    lows = [float(lo) for lo in data["low"]]
    print(f"[INFO] Copper from investing.com ({n} days)")
    return {"dates": dates, "closes": closes, "highs": highs, "lows": lows, "volumes": None, "source": "investing.com"}

def _fetch_ohlc_yfinance():
    """Fallback: fetch COMEX copper OHLCV from yfinance."""
    import yfinance as yf
    ticker = yf.Ticker("HG=F")
    hist = ticker.history(period="1y", interval="1d")
    if hist.empty:
        return None
    hist = hist.reset_index()
    hist.columns = [c if isinstance(c, str) else c[0] for c in hist.columns]
    dates = [r["Date"].to_pydatetime() if hasattr(r["Date"], "to_pydatetime") else r["Date"] for _, r in hist.iterrows()]
    closes = hist["Close"].astype(float).tolist()
    highs = hist["High"].astype(float).tolist()
    lows = hist["Low"].astype(float).tolist()
    volumes = hist["Volume"].astype(float).tolist()
    print(f"[INFO] Copper from yfinance ({len(closes)} days)")
    return {"dates": dates, "closes": closes, "highs": highs, "lows": lows, "volumes": volumes, "source": "yfinance"}

def fetch_copper_data():
    global _copper_cache
    now = time.time()
    if _copper_cache["data"] and (now - _copper_cache["timestamp"]) < COPPER_CACHE_TTL:
        return _copper_cache["data"]

    try:
        # Try investing.com first, fall back to yfinance
        ohlc = None
        try:
            ohlc = _fetch_ohlc_investiny()
        except Exception as e:
            print(f"[WARN] investiny error: {e}")
        if not ohlc:
            ohlc = _fetch_ohlc_yfinance()
        if not ohlc:
            return None

        dates = ohlc["dates"]; closes = ohlc["closes"]; highs = ohlc["highs"]; lows = ohlc["lows"]
        copper_source = ohlc["source"]
        n_closes = len(closes)

        price = closes[-1]; prev_close = closes[-2] if n_closes > 1 else price
        change = price - prev_close
        change_pct = (change / prev_close) * 100 if prev_close else 0

        # Determine previous settlement for RT overlay
        # If daily data includes today (partial bar), prev settle = closes[-2]
        # If daily data ends yesterday, prev settle = closes[-1]
        today_date = datetime.now().date()
        last_ohlc_date = dates[-1].date() if hasattr(dates[-1], 'date') else dates[-1]
        if isinstance(last_ohlc_date, datetime):
            last_ohlc_date = last_ohlc_date.date()
        if last_ohlc_date >= today_date and n_closes > 1:
            _daily_prev_settle = closes[-2]
        else:
            _daily_prev_settle = closes[-1]

        ma50 = sum(closes[-50:]) / min(n_closes, 50)
        ma100 = sum(closes[-100:]) / min(n_closes, 100)
        ma200 = sum(closes[-200:]) / min(n_closes, 200)

        # Volume (only available from yfinance)
        volumes = ohlc.get("volumes")
        if volumes:
            avg_vol = sum(volumes[-20:]) / min(len(volumes), 20)
            vol = volumes[-1]
            vol_ratio = vol / avg_vol if avg_vol > 0 else 1.0
        else:
            vol_ratio = 1.0

        recent = closes[-5:] if n_closes >= 5 else closes
        spark = [{"date": d.strftime("%Y-%m-%d") if hasattr(d, "strftime") else str(d)[:10],
                  "close": round(c, 4)} for d, c in zip(dates[-30:], closes[-30:])]

        today_high = highs[-1]; today_low = lows[-1]
        today_range = today_high - today_low

        c30 = closes[-30:] if n_closes >= 30 else closes
        c90 = closes[-90:] if n_closes >= 90 else closes
        pct_30d = round(sum(1 for c in c30 if c < price) / len(c30) * 100, 0)
        pct_90d = round(sum(1 for c in c90 if c < price) / len(c90) * 100, 0)
        range_30d_low = min(c30); range_30d_high = max(c30)
        range_90d_low = min(c90); range_90d_high = max(c90)

        roc = {}
        for n in [1, 3, 5, 10]:
            if n_closes > n:
                rc = price - closes[-(n+1)]
                roc[f"{n}d"] = {"change": round(rc, 4), "pct": round(rc / closes[-(n+1)] * 100, 2)}

        streak = 0; streak_dir = None
        for i in range(len(closes)-1, 0, -1):
            if closes[i] > closes[i-1]:
                if streak_dir == "up" or streak_dir is None: streak += 1; streak_dir = "up"
                else: break
            elif closes[i] < closes[i-1]:
                if streak_dir == "down" or streak_dir is None: streak += 1; streak_dir = "down"
                else: break
            else: break

        last10_ranges = [highs[i] - lows[i] for i in range(-min(10, n_closes), 0)]
        avg_daily_range = round(sum(last10_ranges) / len(last10_ranges), 4) if last10_ranges else 0
        vol_vs_avg = round(today_range / avg_daily_range, 2) if avg_daily_range > 0 else 1.0

        sr = calc_support_resistance(closes, highs, lows)
        lme = fetch_lme_price()
        lme_price = lme["price_lb"]; lme_mt = lme["price_mt"]; lme_source = lme["source"]
        spread = round(price - lme_price, 4) if lme_price else None
        spread_pct = round((spread / lme_price) * 100, 2) if lme_price and spread else None
        spread_intel = None
        lme_change = None; lme_change_pct = None
        if spread is not None and lme_price:
            history = save_spread_entry(round(price, 4), round(lme_price, 4), round(spread, 4))
            spread_intel = compute_spread_intelligence(history, spread)
            # LME daily change from previous day's spread history entry
            if len(history) >= 2:
                prev_lme = history[-2].get("lme")
                if prev_lme:
                    lme_change = round(lme_price - prev_lme, 4)
                    lme_change_pct = round((lme_change / prev_lme) * 100, 2)

        dxy = fetch_dxy()
        china = get_china_status()
        lme_status = get_lme_status()
        fed = fetch_fed_data()
        warehouse = get_warehouse_data()

        result = {
            "price": round(price, 4), "prev_close": round(prev_close, 4),
            "change": round(change, 4), "change_pct": round(change_pct, 2),
            "ma50": round(ma50, 4), "ma100": round(ma100, 4), "ma200": round(ma200, 4),
            "vol_ratio": round(vol_ratio, 2), "recent_closes": [round(c, 4) for c in recent],
            "sparkline": spark, "copper_source": copper_source,
            "lme_price_lb": lme_price, "lme_price_mt": lme_mt, "lme_source": lme_source,
            "lme_change": lme_change, "lme_change_pct": lme_change_pct,
            "comex_lme_spread": spread, "comex_lme_spread_pct": spread_pct, "spread_intel": spread_intel,
            "today_high": round(today_high, 4), "today_low": round(today_low, 4),
            "today_range": round(today_range, 4),
            "pct_30d": pct_30d, "pct_90d": pct_90d,
            "range_30d": [round(range_30d_low, 4), round(range_30d_high, 4)],
            "range_90d": [round(range_90d_low, 4), round(range_90d_high, 4)],
            "roc": roc, "streak": streak, "streak_dir": streak_dir,
            "avg_daily_range": avg_daily_range, "vol_vs_avg": vol_vs_avg,
            "support_resistance": sr, "dxy": dxy, "china": china, "lme": lme_status,
            "fed": fed, "warehouse": warehouse,
            "_daily_prev_settle": round(_daily_prev_settle, 4),
            "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        }
        _copper_cache = {"data": result, "timestamp": time.time()}
        return result
    except Exception as e:
        print(f"[ERROR] fetch_copper: {e}")
        import traceback; traceback.print_exc()
        return None


# ---------------------------------------------------------------------------
# SIGNALS
# ---------------------------------------------------------------------------
def compute_signals(md):
    if not md: return None
    p = md["price"]; ma50 = md["ma50"]; ma100 = md["ma100"]; ma200 = md["ma200"]
    vr = md["vol_ratio"]; ch = md["change"]; recent = md["recent_closes"]
    above50 = p > ma50; above100 = p > ma100; above200 = p > ma200

    if above50 and above100 and above200: trend, ts = "UPTREND", "strong"
    elif above200 and (above50 or above100): trend, ts = "UPTREND", "moderate"
    elif above200: trend, ts = "NEUTRAL", "weakening"
    elif not above200 and not above100: trend, ts = "DOWNTREND", "strong"
    else: trend, ts = "NEUTRAL", "mixed"

    sd = ch < -CFG["BIG_MOVE"]; md_drop = ch < -CFG["ATTENTION_MOVE"]
    sr = ch > CFG["BIG_MOVE"]; mr = ch > CFG["ATTENTION_MOVE"]
    hv = vr > 2.0; lv = vr < 0.7

    if sd and hv: mt, mdesc = "LIQUIDATION", "High-volume selloff"
    elif sd and lv: mt, mdesc = "FLASH_CRASH", "Sharp drop on thin volume"
    elif sd: mt, mdesc = "BIG_DROP", f"Down {abs(ch):.2f}"
    elif md_drop: mt, mdesc = "DIP", f"Down {abs(ch):.2f}"
    elif sr and hv: mt, mdesc = "BREAKOUT", "Strong rally on high volume"
    elif sr: mt, mdesc = "BIG_RALLY", f"Up {ch:.2f}"
    elif mr: mt, mdesc = "RALLY", f"Up {ch:.2f}"
    else: mt, mdesc = "NORMAL", "Normal trading range"

    cb100 = sum(1 for c in recent if c < ma100)
    cb200 = sum(1 for c in recent if c < ma200)
    tbw = None
    if cb200 >= 3: tbw = "3+ closes below 200DMA \u2014 mills will likely cut bids"
    elif cb100 >= 2: tbw = "2+ closes below 100DMA \u2014 mills may shade bids"

    ft = CFG["FIX_TARGET"]
    if mt in ("LIQUIDATION", "FLASH_CRASH"):
        sig, sc, sd_txt = "BUY OPP", "blue", "Competitors scared \u2014 strong buying opportunity"
    elif mt == "BIG_DROP":
        sig, sc, sd_txt = "BUY OPP", "blue", "Big drop \u2014 buying opportunity"
    elif mt == "DIP":
        sig, sc, sd_txt = "OPPORTUNISTIC", "blue", "Dip day \u2014 lean into buys"
    elif p >= ft:
        sig, sc, sd_txt = "FIX ALERT", "green", f"Above ${ft:.2f} target \u2014 price against unpriced longs"
    elif p >= ft - 0.10:
        sig, sc, sd_txt = "APPROACHING", "yellow", f"${p:.4f} nearing ${ft:.2f} \u2014 get GTCs in place"
    elif trend == "DOWNTREND" and ts == "strong":
        sig, sc, sd_txt = "CAUTION", "orange", "Sustained downtrend \u2014 mills shading bids"
    elif tbw and cb200 >= 3:
        sig, sc, sd_txt = "CAUTION", "orange", tbw
    elif trend == "UPTREND" and ts == "strong":
        sig, sc, sd_txt = "NORMAL", "green", "Strong uptrend \u2014 normal operations"
    else:
        sig, sc, sd_txt = "NORMAL", "green", "Markets stable \u2014 normal operations"

    momentum_note = None
    roc = md.get("roc", {}); stk = md.get("streak", 0); stk_dir = md.get("streak_dir")
    if stk >= 5:
        momentum_note = f"{stk} consecutive {'up' if stk_dir=='up' else 'down'} days"
    elif roc.get("5d") and abs(roc["5d"]["pct"]) > 2:
        d = roc["5d"]
        momentum_note = f"5d: {d['change']:+.2f} ({d['pct']:+.1f}%)"

    spread = md.get("comex_lme_spread"); spread_signal = None
    if spread is not None:
        if spread > 0.15: spread_signal = {"direction": "COMEX PREMIUM", "msg": f"COMEX +{spread:.2f}/lb over LME \u2014 domestic COMEX sales favorable", "color": "green"}
        elif spread > 0.05: spread_signal = {"direction": "COMEX SLIGHT", "msg": f"COMEX +{spread:.2f}/lb \u2014 slight premium", "color": "green"}
        elif spread > -0.05: spread_signal = {"direction": "PARITY", "msg": f"COMEX-LME near parity ({spread:+.2f})", "color": "yellow"}
        elif spread > -0.15: spread_signal = {"direction": "LME PREMIUM", "msg": f"LME +{abs(spread):.2f}/lb \u2014 LME export slightly favorable", "color": "blue"}
        else: spread_signal = {"direction": "LME PREMIUM", "msg": f"LME +{abs(spread):.2f}/lb \u2014 lean into LME exports", "color": "blue"}

    si = md.get("spread_intel")
    if si and spread_signal:
        if si.get("streak") and si["streak"] >= 3:
            spread_signal["msg"] += f" ({si['streak_direction']} {si['streak']}d)"
        if si.get("pct_30d") is not None:
            if si["pct_30d"] > 85: spread_signal["msg"] += " \u2014 at 30d highs"
            elif si["pct_30d"] < 15: spread_signal["msg"] += " \u2014 at 30d lows"

    dxy = md.get("dxy", {}); dxy_signal = None
    if dxy.get("price"):
        dch = dxy.get("change_pct", 0)
        if dch > 0.3: dxy_signal = {"direction": "up", "msg": f"Dollar +{dch:.1f}% \u2014 bearish copper", "color": "red"}
        elif dch < -0.3: dxy_signal = {"direction": "down", "msg": f"Dollar {dch:.1f}% \u2014 tailwind for copper", "color": "green"}
        else: dxy_signal = {"direction": "flat", "msg": f"Dollar flat ({dch:+.1f}%)", "color": "yellow"}

    return {
        "trend": trend, "trend_strength": ts,
        "above_50": above50, "above_100": above100, "above_200": above200,
        "move_type": mt, "move_desc": mdesc,
        "signal": sig, "signal_color": sc, "signal_detail": sd_txt,
        "trend_break_warning": tbw, "momentum_note": momentum_note,
        "spread_signal": spread_signal, "dxy_signal": dxy_signal,
    }


# ---------------------------------------------------------------------------
# GTC SUGGESTIONS
# ---------------------------------------------------------------------------
def gen_gtc(position, md):
    if not position or not md: return []
    p = md["price"]; net = position.get("net_lbs", 0)
    if net <= 0: return []
    tl = CFG["TRUCKLOAD_LBS"]; loads = int(net / tl)
    levels = [l for l in CFG["GTC_LEVELS"] if l > p - 0.05]
    suggestions = []
    if not levels:
        suggestions.append({"action": "PRICE NOW", "detail": f"Above all targets \u2014 fix {loads} loads ({int(net):,} lbs) at ${p:.4f}", "urgency": "high"})
        return suggestions
    per = max(1, loads // len(levels)); rem = loads
    for lvl in levels:
        if rem <= 0: break
        n = min(per, rem); lbs = n * tl
        if lvl <= p:
            suggestions.append({"action": "FIX NOW", "detail": f"Fix {n} load{'s' if n>1 else ''} ({lbs:,} lbs) at ${lvl:.2f} \u2014 price is here", "urgency": "high", "level": lvl})
        elif lvl <= p + 0.10:
            suggestions.append({"action": "GTC", "detail": f"GTC {n} load{'s' if n>1 else ''} ({lbs:,} lbs) at ${lvl:.2f} \u2014 {(lvl-p)*100:.1f}c away", "urgency": "medium", "level": lvl})
        else:
            suggestions.append({"action": "GTC", "detail": f"GTC {n} load{'s' if n>1 else ''} ({lbs:,} lbs) at ${lvl:.2f} \u2014 {(lvl-p)*100:.1f}c away", "urgency": "low", "level": lvl})
        rem -= n
    return suggestions


def calc_risk(pos, md):
    if not pos or not md: return None
    p = md["price"]; net = pos["net_lbs"]; ac = pos["avg_cost"]; hl = pos.get("hedge_lbs", 0)
    uh = net - hl; mtm = (p - ac) * net
    return {
        "net_lbs": net, "unhedged_lbs": uh, "avg_cost": round(ac, 4),
        "mtm_pl": round(mtm, 2), "risk_per_cent": round(uh * 0.01, 2),
        "risk_per_dime": round(uh * 0.10, 2),
        "hedge_pct": round((hl / net) * 100, 1) if net else 0,
        "priced_sales_lbs": pos.get("priced_sales_lbs", 0),
        "unpriced_sales_lbs": pos.get("unpriced_sales_lbs", 0),
        "priced_sales_avg": pos.get("priced_sales_avg", 0),
        "loads_unpriced": int(net / CFG["TRUCKLOAD_LBS"]) if net > 0 else 0,
        "sales_priced_unshipped_lbs": pos.get("sales_priced_unshipped_lbs", 0),
        "sales_unpriced_shipped_lbs": pos.get("sales_unpriced_shipped_lbs", 0),
        "sales_unpriced_unshipped_lbs": pos.get("sales_unpriced_unshipped_lbs", 0),
        "total_inv_po": pos.get("total_inv_po", 0),
        "inventory_cu_lbs": pos.get("inventory_cu_lbs", 0),
        "po_lbs": pos.get("po_lbs", 0),
        "total_sales_lbs": pos.get("priced_sales_lbs", 0) + pos.get("unpriced_sales_lbs", 0),
        "coverage_pct": round((pos.get("total_inv_po", 0) / (pos.get("priced_sales_lbs", 0) + pos.get("unpriced_sales_lbs", 0))) * 100, 1) if (pos.get("priced_sales_lbs", 0) + pos.get("unpriced_sales_lbs", 0)) > 0 else 0,
        "surplus_deficit_lbs": pos.get("total_inv_po", 0) - (pos.get("priced_sales_lbs", 0) + pos.get("unpriced_sales_lbs", 0)),
        "inv_by_commodity": pos.get("inv_by_commodity", {}),
        "sales_by_commodity": pos.get("sales_by_commodity", {}),
        "icw_cu_lbs": pos.get("icw_cu_lbs", 0),
        "chops_solid_lbs": pos.get("chops_solid_lbs", 0),
    }


def calc_margin_projection(pos, md, risk):
    if not pos or not md or not risk: return None
    comex = md.get("price", 0)
    lme = md.get("lme_price_lb", 0)
    avg_cost = risk.get("avg_cost", 0)
    if not comex or not avg_cost: return None

    # Priced sales — use pre-computed aggregates (priced_tons × final_price)
    priced_lbs = pos.get("priced_sales_lbs", 0)
    priced_avg = pos.get("priced_sales_avg", 0)
    priced_rev = priced_lbs * priced_avg

    # Unpriced sales — project each at current market using open_lbs
    unpriced_rev = 0; unpriced_lbs = 0
    for sale_list in [pos.get("sales_unpriced_shipped", []), pos.get("sales_unpriced_unshipped", [])]:
        for sale in sale_list:
            lbs = sale.get("open_lbs", sale.get("lbs", 0))
            spread = sale.get("spread", 0)
            if lbs <= 0: continue
            basis = sale.get("basis", "COMEX")
            if basis == "LME" and lme:
                proj = lme * spread if spread else lme
            else:
                proj = comex - spread if spread else comex
            unpriced_rev += lbs * proj; unpriced_lbs += lbs

    total_lbs = priced_lbs + unpriced_lbs
    total_rev = priced_rev + unpriced_rev
    if total_lbs <= 0: return None

    avg_sell = total_rev / total_lbs
    gm_per_lb = avg_sell - avg_cost
    return {
        "priced_revenue": round(priced_rev, 2),
        "unpriced_revenue_now": round(unpriced_rev, 2),
        "total_revenue_now": round(total_rev, 2),
        "priced_lbs": round(priced_lbs),
        "unpriced_lbs": round(unpriced_lbs),
        "total_lbs": round(total_lbs),
        "avg_sell_price": round(avg_sell, 4),
        "avg_cost": round(avg_cost, 4),
        "gross_margin_per_lb": round(gm_per_lb, 4),
        "total_gross_margin": round(gm_per_lb * total_lbs, 2),
        "margin_pct": round((gm_per_lb / avg_sell) * 100, 2) if avg_sell else 0,
        "comex_now": round(comex, 4),
        "lme_now": round(lme, 4) if lme else None,
    }


def gen_decisions(sig, risk, md, fix_window, roll=None):
    dec = []
    if not sig: return ["Unable to fetch market data"]
    p = md["price"] if md else 0; ft = CFG["FIX_TARGET"]

    # Fix window headline
    if fix_window:
        fw = fix_window
        if fw["score"] >= 65:
            factors_str = " + ".join(fw["factors"][:3]) if fw["factors"] else ""
            dec.append(f"\U0001F7E2 FIX WINDOW: {fw['label']} ({fw['score']}/100) \u2014 {factors_str}")
        elif fw["score"] <= 30:
            factors_str = " + ".join(fw["factors"][:3]) if fw["factors"] else ""
            dec.append(f"\U0001F534 FIX WINDOW: {fw['label']} ({fw['score']}/100) \u2014 {factors_str}")

    china = md.get("china", {}) if md else {}
    if china.get("thin_liquidity"):
        dec.append(f"\u26A0 {china.get('detail', 'SHFE closed')} \u2014 thin liquidity, watch for flash moves")
    elif china.get("status") == "CLOSED" and china.get("reason") == "Lunar New Year":
        dec.append(china.get("detail", "SHFE closed"))

    if risk and risk.get("sales_unpriced_shipped_lbs", 0) > 0:
        dec.append(f"\u26A0 {risk['sales_unpriced_shipped_lbs']:,.0f} lbs shipped but UNPRICED \u2014 fix these first")

    if sig["move_type"] == "LIQUIDATION":
        dec.append("Liquidation event \u2014 competitors scared. Lean into buys.")
    elif sig["move_type"] == "FLASH_CRASH":
        dec.append("Flash crash on thin volume \u2014 buying window.")
    elif sig["move_type"] == "BIG_DROP":
        dec.append("Big drop \u2014 buying opportunity.")
    elif sig["move_type"] == "DIP":
        dec.append("Dip day \u2014 outbid competitors, build relationships.")

    if p >= ft and risk and risk["net_lbs"] > 0:
        dec.append(f"ABOVE ${ft:.2f} TARGET \u2014 {risk.get('loads_unpriced',0)} loads unpriced. Fix some.")
    elif p >= ft - 0.10:
        dec.append(f"${p:.4f} \u2014 approaching ${ft:.2f}. GTCs in place.")

    sr = md.get("support_resistance", {}) if md else {}
    if sr.get("support"):
        s = sr["support"][0]
        dec.append(f"Support at ${s['level']:.2f} ({s['distance']*100:.0f}c below)")
    if sr.get("resistance"):
        r = sr["resistance"][0]
        dec.append(f"Resistance at ${r['level']:.2f} ({r['distance']*100:.0f}c above)")

    ss = sig.get("spread_signal")
    if ss: dec.append(ss["msg"])
    ds = sig.get("dxy_signal")
    if ds and ds["direction"] != "flat": dec.append(ds["msg"])

    # Fed context
    fed = md.get("fed", {}) if md else {}
    if fed.get("first_cut") and fed["first_cut"] != "None priced":
        dec.append(f"Fed: {fed.get('total_cuts_2026', 0)} cut(s) priced by year-end, first in {fed['first_cut']}")
    elif fed.get("yield_10y"):
        dec.append(f"10Y yield: {fed['yield_10y']}% ({fed.get('yield_10y_change', 0):+.2f})")

    # Warehouse
    wh = md.get("warehouse") if md else None
    if wh and wh.get("mt"):
        arrow = "\u2191" if wh["trend"] == "building" else "\u2193" if wh["trend"] == "drawing" else "\u2192"
        dec.append(f"COMEX warehouse: {wh['mt']:,} MT {arrow} ({wh.get('date','')}) \u2014 {'bearish overhang' if wh['trend']=='building' else 'supply tightening' if wh['trend']=='drawing' else 'stable'}")

    # Contract roll alert
    if roll and roll.get("roll_urgency") in ("critical", "warning"):
        nm_ticker = roll.get("next_month", {}).get("ticker", "next contract")
        dec.append(f"\u26A0 {roll['front_month']['ticker']}: {roll['roll_status']} \u2014 liquidity migrating to {nm_ticker}")

    if sig.get("momentum_note"): dec.append(sig["momentum_note"])

    if risk:
        uh = risk["unhedged_lbs"]; rd = risk["risk_per_dime"]
        if uh > 0: dec.append(f"Long {uh:,.0f} lbs unpriced \u2014 ${abs(rd):,.0f} per 10c move")

    # Baseline deviation alert (1 truckload threshold)
    if risk and risk.get("baseline_deviation") is not None:
        dev = risk["baseline_deviation"]
        baseline = risk["baseline_lbs"]
        tl = CFG["TRUCKLOAD_LBS"]
        if abs(dev) > tl:
            if dev > 0:
                dec.append(f"\u26A0 {abs(dev):,.0f} lbs OVER baseline ({baseline:,.0f}) \u2014 consider pricing/trimming")
            else:
                dec.append(f"\u26A0 {abs(dev):,.0f} lbs UNDER baseline ({baseline:,.0f}) \u2014 look for buys")

    # Sales pipeline alerts — flag grades under 2 months of sales
    if risk:
        mf = CFG.get("MONTHLY_FLOW", {})
        sc = risk.get("sales_by_commodity", {})
        thin = []
        for grade, flow in mf.items():
            if flow > 0:
                sal = sc.get(grade, 0)
                months = sal / flow
                if months < 1:
                    thin.append(f"{grade} ({months:.1f}mo)")
                elif months < 2:
                    thin.append(f"{grade} ({months:.1f}mo)")
        if thin:
            dec.append(f"\u26A0 Sales pipeline thin: {', '.join(thin)} \u2014 need to sell")

    if not dec: dec.append("Markets stable \u2014 normal operations")
    return dec


def load_position():
    hf = find_latest_hedge_file()
    if hf:
        print(f"[INFO] Reading: {os.path.basename(hf)}")
        pos = read_hedge_spreadsheet(hf)
        if pos: return pos
    if POSITION_CSV.exists():
        try:
            with open(POSITION_CSV, "r") as f:
                rows = list(csv.DictReader(f))
                if rows:
                    r = rows[0]
                    return {"net_lbs": float(r.get("net_copper_lbs", 0)), "avg_cost": float(r.get("avg_cost_per_lb", 0)),
                            "hedge_lbs": float(r.get("hedge_lbs", 0)), "updated": r.get("date", "unknown"),
                            "source_file": "geomet_position.csv"}
        except Exception as e: print(f"[ERROR] CSV: {e}")
    return None


# ---------------------------------------------------------------------------
# AUTH — simple session-based login
# ---------------------------------------------------------------------------
_USERS = {
    "richard": hashlib.sha256(b"geomet").hexdigest(),
    "jorge": hashlib.sha256(b"geomet").hexdigest(),
}
_sessions = {}  # token -> {"user": ..., "created": timestamp}
SESSION_MAX_AGE = 86400 * 7  # 7 days

def _check_session(cookie_header):
    if not cookie_header: return None
    c = SimpleCookie()
    c.load(cookie_header)
    if "session" not in c: return None
    token = c["session"].value
    s = _sessions.get(token)
    if s and (time.time() - s["created"]) < SESSION_MAX_AGE:
        return s["user"]
    _sessions.pop(token, None)
    return None

def _create_session(user):
    token = secrets.token_hex(32)
    _sessions[token] = {"user": user, "created": time.time()}
    return token

LOGIN_PAGE = """<!DOCTYPE html><html><head><meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<title>Geomet — Login</title>
<style>
@import url('https://fonts.googleapis.com/css2?family=JetBrains+Mono:wght@400;600;700&family=IBM+Plex+Sans:wght@300;400;500;600&display=swap');
*{margin:0;padding:0;box-sizing:border-box}
body{font-family:'IBM Plex Sans',sans-serif;background:#0a0e14;color:#e0e6ed;min-height:100vh;display:flex;align-items:center;justify-content:center}
.box{background:#111820;border:1px solid #1e2a3a;border-radius:8px;padding:40px;width:320px}
.logo{font-family:'JetBrains Mono',monospace;font-size:11px;font-weight:700;letter-spacing:3px;color:#d4845a;padding:4px 8px;border:1.5px solid #d4845a;border-radius:3px;display:inline-block;margin-bottom:20px}
h2{font-size:14px;font-weight:300;color:#6b7f99;margin-bottom:20px}
input{width:100%;padding:10px 12px;margin-bottom:12px;background:#1a2230;border:1px solid #1e2a3a;border-radius:4px;color:#e0e6ed;font-family:'IBM Plex Sans',sans-serif;font-size:13px}
input:focus{outline:none;border-color:#d4845a}
button{width:100%;padding:10px;background:rgba(212,132,90,.15);border:1px solid #d4845a;border-radius:4px;color:#d4845a;font-family:'JetBrains Mono',monospace;font-size:11px;font-weight:600;letter-spacing:1px;cursor:pointer}
button:hover{background:rgba(212,132,90,.25)}
.err{color:#ff4757;font-size:11px;margin-bottom:10px;display:none}
</style></head><body>
<div class="box"><div class="logo">GEOMET</div><h2>Copper Intelligence Dashboard</h2>
<div class="err" id="err">Invalid username or password</div>
<form method="POST" action="/login">
<input name="user" placeholder="Username" autocomplete="username" required>
<input name="pass" type="password" placeholder="Password" autocomplete="current-password" required>
<button type="submit">LOGIN</button></form></div>
<script>if(location.search.includes('err=1'))document.getElementById('err').style.display='block'</script>
</body></html>"""

THEME_FILE = DATA_DIR / "theme.json"

def get_theme():
    if THEME_FILE.exists():
        try:
            with open(THEME_FILE) as f: return json.load(f).get("theme", "dark")
        except: pass
    return "dark"

def set_theme(theme):
    try:
        with open(THEME_FILE, "w") as f: json.dump({"theme": theme}, f)
    except: pass

class Handler(SimpleHTTPRequestHandler):
    def _authed(self):
        return _check_session(self.headers.get("Cookie"))

    def do_GET(self):
        if self.path == "/login":
            self.send_response(200)
            self.send_header("Content-Type", "text/html"); self.end_headers()
            self.wfile.write(LOGIN_PAGE.encode()); return
        if self.path == "/logout":
            c = SimpleCookie()
            c.load(self.headers.get("Cookie") or "")
            if "session" in c: _sessions.pop(c["session"].value, None)
            self.send_response(302)
            self.send_header("Set-Cookie", "session=; Path=/; Max-Age=0")
            self.send_header("Location", "/login"); self.end_headers(); return
        if not self._authed():
            self.send_response(302)
            self.send_header("Location", "/login"); self.end_headers(); return
        if self.path == "/api/theme":
            self.send_response(200)
            self.send_header("Content-Type", "application/json")
            self.send_header("Access-Control-Allow-Origin", "*")
            self.end_headers()
            self.wfile.write(json.dumps({"theme": get_theme()}).encode())
            return
        if self.path == "/api/data":
            self.send_response(200)
            self.send_header("Content-Type", "application/json")
            self.send_header("Access-Control-Allow-Origin", "*")
            self.end_headers()
            md = fetch_copper_data()
            # Overlay real-time price on top of cached daily OHLC
            rt = _fetch_realtime_price()
            if rt.get("price") and md:
                md = dict(md)  # copy so we don't mutate the cache
                prev = rt.get("prev_close") or md.get("_daily_prev_settle", md["prev_close"])
                md["price"] = rt["price"]
                md["prev_close"] = prev
                md["change"] = round(rt["price"] - prev, 4)
                md["change_pct"] = round(((rt["price"] - prev) / prev) * 100, 2) if prev else 0
                md["copper_source"] = rt.get("source", md.get("copper_source", ""))
                # Tell frontend which contract the big price represents
                md["active_contract"] = rt.get("active_contract", "front")
                # Recalculate COMEX-LME spread with RT price
                if md.get("lme_price_lb"):
                    md["comex_lme_spread"] = round(rt["price"] - md["lme_price_lb"], 4)
            sig = compute_signals(md)
            pos = load_position()
            risk = calc_risk(pos, md)
            if risk:
                risk["baseline_lbs"] = CFG["BASELINE_LBS"]
                risk["baseline_deviation"] = risk["net_lbs"] - CFG["BASELINE_LBS"]
            fix_window = calc_fix_window(md, sig)
            # When RT source is the next month (e.g. investing.com → May),
            # pass None as copper_price so contract_roll fetches front independently.
            # Then we override the spread with the correct direction.
            _ac = md.get("active_contract", "front") if md else "front"
            roll = get_contract_roll(md.get("price") if (_ac == "front" and md) else None)
            if roll and _ac == "next" and md:
                # Big price is May (next). Use it as next_price, fetch front separately.
                roll["next_price"] = md["price"]
                if roll.get("front_price"):
                    spread = round(md["price"] - roll["front_price"], 4)
                    roll["calendar_spread"] = spread
                    roll["market_structure"] = "contango" if spread > 0.001 else "backwardation" if spread < -0.001 else "flat"
            dec = gen_decisions(sig, risk, md, fix_window, roll)
            gtc = gen_gtc(pos, md)
            margin = calc_margin_projection(pos, md, risk)
            fixable = calc_fixable_orders(pos, md)
            payload = {
                "market": md, "signals": sig, "position": pos, "position_risk": risk,
                "decisions": dec, "gtc_suggestions": gtc, "fix_window": fix_window,
                "fixable_orders": fixable,
                "margin_projection": margin, "contract_roll": roll,
                "config": {"fix_target": CFG["FIX_TARGET"], "truckload_lbs": CFG["TRUCKLOAD_LBS"], "gtc_levels": CFG["GTC_LEVELS"], "baseline_lbs": CFG["BASELINE_LBS"], "monthly_flow": CFG["MONTHLY_FLOW"]},
                "last_refresh": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            }
            self.wfile.write(json.dumps(payload).encode())
            return
        if self.path in ("/", ""): self.path = "/index.html"
        fp = STATIC_DIR / self.path.lstrip("/")
        if fp.exists() and fp.is_file():
            self.send_response(200)
            ct = "text/html" if str(fp).endswith(".html") else "text/css" if str(fp).endswith(".css") else "application/javascript"
            self.send_header("Content-Type", ct); self.end_headers()
            self.wfile.write(fp.read_bytes())
        else: self.send_error(404)
    def do_POST(self):
        if self.path == "/login":
            length = int(self.headers.get("Content-Length", 0))
            body = self.rfile.read(length).decode()
            from urllib.parse import parse_qs
            params = parse_qs(body)
            user = params.get("user", [""])[0].strip().lower()
            pwd = params.get("pass", [""])[0]
            pwd_hash = hashlib.sha256(pwd.encode()).hexdigest()
            if user in _USERS and _USERS[user] == pwd_hash:
                token = _create_session(user)
                self.send_response(302)
                self.send_header("Set-Cookie", f"session={token}; Path=/; Max-Age={SESSION_MAX_AGE}; HttpOnly; SameSite=Lax")
                self.send_header("Location", "/"); self.end_headers()
            else:
                self.send_response(302)
                self.send_header("Location", "/login?err=1"); self.end_headers()
            return
        if not self._authed():
            self.send_response(403)
            self.send_header("Content-Type", "application/json"); self.end_headers()
            self.wfile.write(b'{"error":"unauthorized"}'); return
        if self.path == "/api/theme":
            length = int(self.headers.get("Content-Length", 0))
            body = json.loads(self.rfile.read(length)) if length else {}
            theme = body.get("theme", "dark")
            if theme not in ("dark", "light"): theme = "dark"
            set_theme(theme)
            self.send_response(200)
            self.send_header("Content-Type", "application/json")
            self.send_header("Access-Control-Allow-Origin", "*")
            self.end_headers()
            self.wfile.write(json.dumps({"theme": theme}).encode())
            return
        self.send_error(404)
    def log_message(self, *a): pass

def main():
    DATA_DIR.mkdir(exist_ok=True); STATIC_DIR.mkdir(exist_ok=True)
    hf = find_latest_hedge_file()
    lme_src = "metals.dev API" if CFG["METALS_DEV_API_KEY"] else ("manual" if CFG["LME_MANUAL_USD_MT"] else "NOT CONFIGURED")
    print()
    print("=" * 55)
    print("  GEOMET COPPER INTELLIGENCE DASHBOARD v4.1")
    print("=" * 55)
    print(f"  Dashboard:  http://localhost:{PORT}")
    print(f"  Fix target: ${CFG['FIX_TARGET']}")
    print(f"  LME source: {lme_src}")
    print(f"  LME cache:  30 min (market hours only)")
    fed_src = "FRED API" if CFG["FRED_API_KEY"] else "static config"
    print(f"  Fed rate:   {CFG['FED_FUNDS_RATE']}% ({fed_src})")
    wh = CFG.get("COMEX_WAREHOUSE_MT", 0)
    if wh: print(f"  Warehouse:  {wh:,} MT ({CFG.get('COMEX_WAREHOUSE_DATE','')})")
    if hf: print(f"  Hedge file: {os.path.basename(hf)}")
    print("=" * 55)
    print()
    socketserver.TCPServer.allow_reuse_address = True
    with socketserver.TCPServer(("", PORT), Handler) as httpd:
        httpd.serve_forever()

if __name__ == "__main__":
    main()
